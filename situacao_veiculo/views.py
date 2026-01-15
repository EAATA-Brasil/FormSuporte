from datetime import date, datetime

from django.db import IntegrityError
from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_GET, require_POST
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
import requests
from email.utils import parsedate_to_datetime

from .models import Cliente
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
from ocorrencia_erro.models import Device

def buscar_serial(request):
    context = {}
    if request.method == 'POST':
        serial = request.POST.get('serial', '').strip()
        context['serial_digitado'] = serial

        # Busca case-insensitive para evitar falhas por maiúsculas/minúsculas
        clientes = Cliente.objects.filter(serial__iexact=serial)

        if not clientes.exists():
            # Busca em serviço externo usando o serial digitado (headers e cookie conforme curl)
            try:
                cookie_value = 'eyJ1c2VyX2lkIjoiZWFhdGFkbWluIn0.aV_2TA.URbvR1qfUJFd6H56IRWZc_hSSp0'
                headers = {
                    'Accept': '*/*',
                    'Accept-Language': 'pt-PT,pt;q=0.9,en-US;q=0.8,en;q=0.7,es;q=0.6',
                    'Connection': 'keep-alive',
                    'Content-Type': 'application/json',
                    'Origin': 'http://20.83.150.13:8088',
                    'Referer': 'http://20.83.150.13:8088/dashboard',
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36',
                }
                resp = requests.post(
                    'http://20.83.150.13:8088/search_codes',
                    json={'sn': serial},
                    headers=headers,
                    cookies={'session': cookie_value},
                    timeout=10,
                )
                print(resp.text)
                try:
                    resp_json = resp.json()
                except ValueError:
                    resp_json = {'status_code': resp.status_code, 'text': resp.text}

                # Prepare a rich external_search payload for the template/front
                external_info = {'status': resp.status_code, 'data': resp_json}
                print(external_info)
                try:
                    if isinstance(resp_json, dict):
                        codes = resp_json.get('codes')
                        if codes and isinstance(codes, list) and len(codes) > 0:
                            first = codes[0]
                            created_at_str = first.get('created_at')
                            email = first.get('email')
                            external_info['cliente'] = email
                            external_info['sn'] = first.get('sn')

                            if created_at_str:
                                try:
                                    created_dt = parsedate_to_datetime(created_at_str)
                                    # Compute +2 years (fallback for Feb 29)
                                    try:
                                        venc_dt = created_dt.replace(year=created_dt.year + 2)
                                    except ValueError:
                                        venc_dt = created_dt.replace(month=2, day=28, year=created_dt.year + 2)
                                    external_info['created_at'] = created_dt.isoformat()
                                    external_info['vencimento'] = venc_dt.date().isoformat()
                                except Exception:
                                    pass
                except Exception:
                    pass

                context['external_search'] = external_info
            except requests.RequestException as exc:
                context['external_search'] = {'error': str(exc)}

            # Se o serviço externo retornou algum código, sinaliza que é preciso atualizar os dados
            found_external = False
            es = context.get('external_search')
            if isinstance(es, dict):
                data = es.get('data')
                if isinstance(data, dict):
                    codes = data.get('codes')
                    if codes and isinstance(codes, list) and len(codes) > 0:
                        found_external = True

            if found_external:
                # Indica que os dados foram obtidos externamente e precisam de atualização
                context['mensagem'] = 'Dados captaados externamente, necessário atualização'
                # Mapear dados externos para o formato que o template espera (cliente, vencimento)
                data = es.get('data') if isinstance(es, dict) else None
                first = None
                if isinstance(data, dict):
                    codes = data.get('codes')
                    if codes and isinstance(codes, list) and len(codes) > 0:
                        first = codes[0]

                cliente_dict = {
                    'nome': es.get('email') or (first.get('email') if first else '') or '',
                    'cnpj': '',
                    'tel': '',
                    'equipamento': ('{} - {}'.format(first.get('city',''), first.get('country','')).strip(' -') if first else ''),
                    'vencimento': es.get('vencimento') or (first.get('created_at') if first else None),
                }
                context['cliente'] = cliente_dict

                # Determina status com base em vencimento (created_at + 2 anos)
                status_val = 'indefinido'
                venc_str = es.get('vencimento') if isinstance(es, dict) else None
                if venc_str:
                    try:
                        venc_date = parse_date(venc_str)
                        if isinstance(venc_date, datetime):
                            venc_date = venc_date.date()
                        dias = (venc_date - date.today()).days
                        if dias > 30:
                            status_val = 'direito'
                        elif dias < 1:
                            status_val = 'vencido'
                        else:
                            status_val = 'vencendo'
                    except Exception:
                        status_val = 'indefinido'
                else:
                    # fallback: calcule a partir de created_at se estiver disponível
                    created_at_str = None
                    if first:
                        created_at_str = first.get('created_at')
                    if created_at_str:
                        try:
                            created_dt = parsedate_to_datetime(created_at_str)
                            try:
                                venc_dt = created_dt.replace(year=created_dt.year + 2)
                            except ValueError:
                                venc_dt = created_dt.replace(month=2, day=28, year=created_dt.year + 2)
                            dias = (venc_dt.date() - date.today()).days
                            if dias > 30:
                                status_val = 'direito'
                            elif dias < 1:
                                status_val = 'vencido'
                            else:
                                status_val = 'vencendo'
                            if not cliente_dict.get('vencimento'):
                                cliente_dict['vencimento'] = venc_dt.date().isoformat()
                        except Exception:
                            status_val = 'indefinido'

                context['status'] = status_val
                # Map status to short message following the same rules as models.status_message_default
                if status_val == 'direito':
                    context['status_message'] = "SUPORTE LIBERADO - Atualizar dados e atender normalmente"
                elif status_val == 'vencido':
                    context['status_message'] = "SUPORTE VENCIDO - Não fazer atendimento"
                elif status_val == 'vencendo':
                    context['status_message'] = "SUPORTE A VENCER - Atualizar dados e atender normalmente"
            else:
                context['status_message'] = 'SEM DADOS'
                context['mensagem'] = 'Passar para o comercial atualizar o cadastro.'

            return render(request, 'situacao/index.html', context)

        if clientes.count() > 1:
            lista_clientes = []
            for cliente in clientes:
                lista_clientes.append({
                    'cliente': cliente,
                    'status': cliente.status,
                    'vencimento_dias': getattr(cliente, '_vencimento_dias', None),
                    'status_message': cliente.status_message,  # já é “efetiva”
                })
            context['clientes_duplicados'] = lista_clientes
            context['mensagem'] = 'Encontradas múltiplas ocorrências para esse serial. Verifique os dados abaixo:'
            return render(request, 'situacao/index.html', context)

        # Único cliente
        cliente = clientes.first()
        context['cliente'] = cliente
        context['status'] = cliente.status
        context['mensagem'] = cliente.message_effective   # detalhada efetiva
        context['status_message'] = cliente.status_message  # curta efetiva
        return render(request, 'situacao/index.html', context)

    return redirect('index')  # GET -> homepage

def _anos_por_equipamento(equipamento: str) -> int:
    if not equipamento:
        return 2
    return 1 if equipamento.lower().find('reader') != -1 else 2


ALLOWED_FIELDS = {"nome", "cnpj", "tel", "vencimento"}

def _digits_only(s: str) -> str:
    return ''.join(ch for ch in s if ch.isdigit())


def _parse_excel_date(value, workbook):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value, epoch=workbook.epoch).date()
        except Exception:
            return None
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        parsed = parse_date(cleaned)
        if parsed:
            return parsed
        try:
            return datetime.strptime(cleaned, "%d/%m/%Y").date()
        except ValueError:
            return None
    return None


@require_POST
def cadastrar_serial(request):
    data_referencia = timezone.now()

    serial = (request.POST.get('serial') or '').strip()
    nome = (request.POST.get('nome') or '').strip()
    cnpj = (request.POST.get('cnpj') or '').strip()
    tel = (request.POST.get('tel') or '').strip()
    equipamento = (request.POST.get('equipamento') or '').strip()
    data_input = (request.POST.get('data') or '').strip()

    # Novos (somente para superuser)
    anos_input = (request.POST.get('anos_para_vencimento') or '').strip()
    venc_input = (request.POST.get('vencimento') or '').strip()

    field_errors = {}
    if not serial:
        field_errors['serial'] = 'Serial é obrigatório.'

    # Validações dos novos campos (apenas se superuser enviou algo)
    anos_para_vencimento = None
    vencimento_data = None

    if request.user.is_superuser:
        # anos_para_vencimento: opcional, default 2 se não vier
        if anos_input:
            try:
                anos_para_vencimento = int(anos_input)
                if anos_para_vencimento < 0:
                    field_errors['anos_para_vencimento'] = 'Deve ser um inteiro zero ou positivo.'
            except ValueError:
                field_errors['anos_para_vencimento'] = 'Informe um número inteiro.'
        else:
            anos_para_vencimento = 2  # default solicitado

        # vencimento: opcional (pode vir vazio), valida formato se vier
        if venc_input:
            d = parse_date(venc_input)
            if not d:
                field_errors['vencimento'] = 'Data inválida (use YYYY-MM-DD).'
            else:
                vencimento_data = d

    if field_errors:
        return JsonResponse(
            {"ok": False, "message": "Corrija os campos destacados.", "field_errors": field_errors},
            status=400,
        )

    try:
        # Evita duplicidade ignorando caixa
        if Cliente.objects.filter(serial__iexact=serial).exists():
            return JsonResponse(
                {"ok": False, "message": "Serial já em uso.", "field_errors": {"serial": "Serial já cadastrado."}},
                status=409,
            )

        # Se não for superuser, mantém sua lógica atual (por equipamento)
        if not request.user.is_superuser:
            anos_para_vencimento = _anos_por_equipamento(equipamento)


        data_lanc = None
        if data_input:
            d = parse_date(data_input)
            if not d:
                field_errors['data'] = 'Data inválida (use YYYY-MM-DD).'
            else:
                data_lanc = d


        cliente = Cliente.objects.create(
            data=data_lanc or timezone.now(),
            anos_para_vencimento=int(anos_para_vencimento),
            serial=serial,
            nome=nome,
            cnpj=_digits_only(cnpj) or cnpj,
            tel=tel,
            equipamento=equipamento or "N/D",
            # vencimento: só setamos se superuser enviou; senão, o models.save() cuidará (auto) quando None
            vencimento=vencimento_data if request.user.is_superuser else None,
        )

        return JsonResponse(
            {
                "ok": True,
                "message": "Cadastro realizado com sucesso!",
                "data": {
                    "id": cliente.id,
                    "serial": cliente.serial,
                    "nome": cliente.nome,
                    "cnpj": cliente.cnpj,
                    "tel": cliente.tel,
                    "equipamento": cliente.equipamento,
                    "anos_para_vencimento": cliente.anos_para_vencimento,
                    "vencimento": cliente.vencimento.isoformat() if cliente.vencimento else None,
                    "timestamp": data_referencia.isoformat(),
                },
            },
            status=201,
        )

    except IntegrityError:
        return JsonResponse(
            {"ok": False, "message": "Não foi possível cadastrar (restrição de unicidade)."},
            status=409,
        )
    except ValueError as e:
        return JsonResponse(
            {"ok": False, "message": f"Erro de validação: {e}"},
            status=400,
        )
    except Exception:
        return JsonResponse(
            {"ok": False, "message": "Erro interno ao cadastrar."},
            status=500,
        )

@require_GET
def api_buscar_cliente(request):
    serial = (request.GET.get('serial') or '').strip()
    if not serial:
        return JsonResponse({"ok": False, "message": "Informe o serial."}, status=400)

    try:
        cliente = Cliente.objects.filter(serial__iexact=serial).first()
    except Cliente.DoesNotExist:
        cliente = None
    if not cliente:
        return JsonResponse({"ok": False, "message": "Serial não encontrado."}, status=404)

    data = {
        "id": cliente.id,
        "serial": cliente.serial,
        "nome": cliente.nome,
        "cnpj": cliente.cnpj,
        "tel": cliente.tel,
        "vencimento": cliente.vencimento.isoformat() if cliente.vencimento else None,
    }
    return JsonResponse({"ok": True, "data": data}, status=200)


@require_POST
def api_atualizar_cliente(request):
    serial = (request.POST.get('serial') or '').strip()
    field = (request.POST.get('field') or '').strip()
    value = (request.POST.get('value') or '')

    if not serial:
        return JsonResponse({"ok": False, "message": "Serial é obrigatório."}, status=400)
    if field not in ALLOWED_FIELDS:
        return JsonResponse({"ok": False, "message": "Campo não permitido para atualização."}, status=400)

    cliente = Cliente.objects.filter(serial__iexact=serial).first()
    if not cliente:
        return JsonResponse({"ok": False, "message": "Serial não encontrado."}, status=404)

    # Normalizações específicas
    if field == "cnpj":
        value = ''.join(ch for ch in value if ch.isdigit()) or value

    if field == "vencimento":
        if not value:
            cliente.vencimento = None
            cliente.save(update_fields=["vencimento"])
            return JsonResponse({"ok": True, "message": "Vencimento removido."})
        data_v = parse_date(value)
        if not data_v:
            return JsonResponse({"ok": False, "message": "Data inválida. Use formato YYYY-MM-DD."}, status=400)
        cliente.vencimento = data_v
        cliente.save(update_fields=["vencimento"])
        return JsonResponse({"ok": True, "message": "Vencimento atualizado.", "data": {"vencimento": data_v.isoformat()}})

    setattr(cliente, field, value)
    cliente.save(update_fields=[field])

    return JsonResponse({"ok": True, "message": f"{field.capitalize()} atualizado.", "data": {field: value}})


@require_POST
def importar_excel(request):
    planilha = request.FILES.get('arquivo_excel')
    if not planilha:
        return JsonResponse({"ok": False, "message": "Envie um arquivo Excel."}, status=400)

    try:
        workbook = load_workbook(planilha, data_only=True)
    except Exception:
        return JsonResponse({"ok": False, "message": "Não foi possível ler o arquivo enviado."}, status=400)

    sheet = workbook.active
    header_cells = sheet[1]

    expected = {
        "nome cliente": "nome",
        "nome item": "equipamento",
        "serial": "serial",
        "cnpj/cpf": "cnpj",
        "contato": "tel",
        "numero emissão nf": "data",
    }

    header_map = {}
    for idx, cell in enumerate(header_cells):
        header = str(cell.value).strip().lower() if cell.value else ""
        if header in expected:
            header_map[expected[header]] = idx

    if "serial" not in header_map:
        return JsonResponse(
            {"ok": False, "message": "A coluna 'serial' é obrigatória na planilha."},
            status=400,
        )

    created = 0
    duplicates = 0
    errors = []

    for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        row_data = {}
        for field, col_idx in header_map.items():
            value = row[col_idx].value if col_idx < len(row) else None
            row_data[field] = value

        if not any(row_data.values()):
            continue

        serial = (row_data.get("serial") or "").strip()
        if not serial:
            errors.append({"row": row_index, "message": "Serial ausente."})
            continue

        if Cliente.objects.filter(serial__iexact=serial).exists():
            duplicates += 1
            continue

        nome = (row_data.get("nome") or "").strip()
        equipamento = (row_data.get("equipamento") or "").strip() or "N/D"
        cnpj = _digits_only((row_data.get("cnpj") or "").strip())
        tel = (row_data.get("tel") or "").strip()

        data_valor = _parse_excel_date(row_data.get("data"), workbook)
        data_lanc = data_valor or timezone.localdate()

        try:
            Cliente.objects.create(
                data=data_lanc,
                anos_para_vencimento=_anos_por_equipamento(equipamento),
                serial=serial,
                nome=nome,
                cnpj=cnpj or None,
                tel=tel or None,
                equipamento=equipamento,
            )
            created += 1
        except IntegrityError:
            duplicates += 1
        except Exception as exc:
            errors.append({"row": row_index, "message": str(exc)})

    message = "Importação concluída."
    if created or duplicates or errors:
        message = (
            f"Importação concluída. Criados: {created}. Duplicados ignorados: {duplicates}. "
            f"Erros: {len(errors)}."
        )

    status_code = 200 if not errors else 207
    return JsonResponse(
        {"ok": True, "message": message, "data": {"created": created, "duplicates": duplicates, "errors": errors}},
        status=status_code,
    )


@require_GET
def equipamentos_suggest(request):
    """Autocomplete de equipamentos baseado na tabela Device (ocorrencia_erro).
    Parâmetros:
      - q: trecho a buscar (case-insensitive). Opcional; sem q retorna os primeiros.
      - limit: máximo de resultados (default 15, máx 50)
    Retorna: { results: ["nome1", "nome2", ...] }
    """
    q = (request.GET.get('q') or '').strip()
    try:
        limit = int(request.GET.get('limit') or 15)
    except ValueError:
        limit = 15
    limit = max(1, min(limit, 50))

    qs = Device.objects.all()
    if q:
        qs = qs.filter(name__icontains=q)
    names = list(qs.order_by('name').values_list('name', flat=True)[:limit])
    return JsonResponse({"results": names})


def index(request):
    return render(request, 'situacao/index.html', {'clientes': None})
